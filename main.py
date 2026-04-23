import os
import pandas as pd
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import load_workbook
import glob
#script_dir = os.path.dirname(os.path.abspath(__C:\Users\mctrainer\Desktop\survey_visualization\survey_visualization__))

# database connection 
engine = create_engine("mysql+pymysql://root:root@localhost/survey_db")

CONFIG_FILE = "survey_config_template.xlsx"
#RAW_FILE = "2024Q2_RS_0613_0629.xlsx"

# normalize text for matching 
def normalize(text):
    if pd.isna(text):
        return ""
    return str(text).strip().lower().replace(".", "")

# read config data and prepare mappings
def load_config():
    question_master = pd.read_excel(CONFIG_FILE, sheet_name="question_master")
    column_mapping = pd.read_excel(CONFIG_FILE, sheet_name="column_mapping")
    scale_mapping = pd.read_excel(CONFIG_FILE, sheet_name="scale_mapping")

    column_mapping["pattern_norm"] = column_mapping["raw_column_pattern"].apply(normalize)
    scale_mapping["answer_norm"] = scale_mapping["answer_text"].apply(normalize)

    return question_master, column_mapping, scale_mapping

# read raw data for both surveys
def load_raw_data(RAW_FILE):
    raw_data = pd.read_excel(RAW_FILE, sheet_name=None)
    raw_data_dict = {
        "MODULE": raw_data["Raw Review"],
        "TRAQOM": raw_data["TRAQOM Raw Data"][raw_data["TRAQOM Raw Data"]["_sourceType"] == "survey"]
    }
    return raw_data_dict

# match raw columns to question codes based on config patterns
def match_columns(raw_df, column_mapping, survey_type):
    mapping = column_mapping[column_mapping["survey_type"] == survey_type]
    result = {}
    for col in raw_df.columns:
        col_norm = normalize(col)
        for _, row in mapping.iterrows():
            if row["pattern_norm"] in col_norm:
                result[row["question_code"]] = raw_df[col]
                break  
    matched_df = pd.DataFrame(result)
    return matched_df

# change raw answers to numeric scores using scale mapping
comment_cols = ["M16", "M17", "M18", "T11", "T12","T13"]
def convert_scores(df, scale_mapping):
    score_dict = dict(zip(scale_mapping["answer_norm"], scale_mapping["score"]))
    
    def convert_value(x):
        if pd.isna(x):
            return None
        if isinstance(x, (int, float)):
            return x
        x_norm = normalize(x)
        return score_dict.get(x_norm, None)
    
    # avoid modifying the original DataFrame
    df_copy = df.copy()
    for col in df_copy.columns:
        if col not in comment_cols:
            df_copy[col] = df_copy[col].apply(convert_value)
    return df_copy

# build wide table with metadata columns 
def build_wide_table(score_df, raw_df, survey_type):
    wide_df = score_df.copy()
    wide_df["survey_type"] = survey_type
    
    if survey_type == "MODULE":
        wide_df["course_name"] = None
        wide_df["trainer_name"] = None
        wide_df["course_start"] = None
        wide_df["course_end"] = None
        wide_df["course_run"] = None
    elif survey_type == "TRAQOM":
        wide_df["course_name"] = raw_df["Course Title"]
        wide_df["trainer_name"] = raw_df["Trainer Name 1"]
        wide_df["course_start"] = raw_df["Course Start Date (+08:00 GMT)"]
        wide_df["course_end"] = raw_df["Course End Date (+08:00 GMT)"]
        wide_df["course_run"] = raw_df["Course Run Id"]
    
    # change date format
    def convert_date(date_str):
        if pd.isna(date_str):
            return None
        
        date_str = str(date_str)
    
    # delete +08:00 GMT
        date_str = date_str.replace(" +08:00 GMT", "")

# define all the date format
        formats = [
            "%m/%d/%Y %H:%M",    # 9/10/2025 8:00 TRAQOM Date Columns
            "%m/%d/%Y",          # 6/3/2024
            "%Y-%m-%d %H:%M:%S", # 2025-09-04 00:00:00 
            "%Y-%m-%d"           # 2025-09-04 
        ]

        for fmt in formats:
            try:
                dt = datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue

# if all failed
        print(f"can't read {date_str}")
        return None

    wide_df["course_start"] = wide_df["course_start"].apply(convert_date)
    wide_df["course_end"] = wide_df["course_end"].apply(convert_date)


    meta_cols = ["survey_type", "course_name", "course_run", "trainer_name", "course_start", "course_end"]
    score_cols = list(score_df.columns)
    wide_df = wide_df[meta_cols + score_cols]
    return wide_df

# use TRAQOM data to fill MODULE's missing metadata (add null value checks)
def fill_module_metadata(final_wide):
    traqom_data = final_wide[final_wide["survey_type"] == "TRAQOM"]
    course_name = traqom_data["course_name"].dropna().iloc[0] if not traqom_data["course_name"].dropna().empty else None
    trainer_name = traqom_data["trainer_name"].dropna().iloc[0] if not traqom_data["trainer_name"].dropna().empty else None
    course_start = traqom_data["course_start"].dropna().iloc[0] if not traqom_data["course_start"].dropna().empty else None
    course_end = traqom_data["course_end"].dropna().iloc[0] if not traqom_data["course_end"].dropna().empty else None
    course_run = traqom_data["course_run"].dropna().iloc[0] if not traqom_data["course_run"].dropna().empty else None
    
    # fill in MODULE none
    mask = final_wide["survey_type"] == "MODULE"
    if trainer_name:
        final_wide.loc[mask, "trainer_name"] = trainer_name
    if course_start:
        final_wide.loc[mask, "course_start"] = course_start
    if course_end:
        final_wide.loc[mask, "course_end"] = course_end
    if course_run:
        final_wide.loc[mask, "course_run"] = course_run
    if course_name: 
        final_wide.loc[mask, "course_name"] = course_name
    return final_wide

# build long table (add response_id as id column)
def build_long_table(wide_df):
    id_cols = [
        "response_id",  
        "survey_type",
        "course_name",
        "course_run",
        "trainer_name",
        "course_start",
        "course_end"
    ]

    id_cols = [col for col in id_cols if col in wide_df.columns]
    
    long_df = wide_df.melt(
        id_vars=id_cols,
        var_name="question_code",
        value_name="score"
    )

    comment_df = long_df[long_df["question_code"].isin(comment_cols)].copy()

    comment_df = comment_df.rename(columns={"score": "comment"})

    comment_df = comment_df.dropna(subset=["comment"])
    comment_df = comment_df[comment_df["comment"].str.strip() != ""]

    comment_df = comment_df[["response_id", "question_code", "comment"]]
    comment_df.to_sql(name="survey_comment", con=engine, if_exists="append", index=False)

    long_df = long_df.dropna(subset=["score"])
    return long_df

# main function to orchestrate the process with error handling
def main():
    try:
        Raw_FOLDER = "raw_data"
        all_files = glob.glob(os.path.join(Raw_FOLDER, "*.xlsx"))
        for f in all_files:
            print(f"are processing: {f}")

            print("Loading config...")
            question_master, column_mapping, scale_mapping = load_config()

            print("Loading raw data...")
            raw_data_dict = load_raw_data(f)

            wide_tables = []
            for survey_type, raw_df in raw_data_dict.items():
                print("\nProcessing:", survey_type)
                matched_df = match_columns(raw_df, column_mapping, survey_type)
                score_df = convert_scores(matched_df, scale_mapping)
                wide_df = build_wide_table(score_df, raw_df, survey_type)
                wide_tables.append(wide_df)
                print("Rows processed:", len(wide_df))

            final_wide = pd.concat(wide_tables, ignore_index=True)
            
            final_wide = fill_module_metadata(final_wide)

            # write response_id back to wide table before generating long table
            # course_metadata
            course_df = final_wide[
                ["course_run", "course_name", "trainer_name", "course_start", "course_end"]
            ].drop_duplicates().dropna(subset=["course_run"])  # filter no course_run recordings

            if not course_df.empty:
                course_df.to_sql(
                    "course_metadata",
                    engine,
                    if_exists="append",
                    index=False
                )
                print(f" course_metadata finished {len(course_df)} course records")
            else:
                print(" course_metadata no vaild data to write")

            # survey_response
            response_df = final_wide[["course_run", "survey_type"]].copy()
            
            if not response_df.empty:
                response_df.to_sql(
                    "survey_response",
                    engine,
                    if_exists="append",
                    index=False
                )
                print(f"survey_response finished {len(response_df)} response records")

                # readresponse_id
                ids = pd.read_sql(
                    f"""
                    SELECT response_id
                    FROM survey_response
                    ORDER BY response_id DESC
                    LIMIT {len(response_df)}
                    """,
                    engine
                )
                # reverse ids to match the original order of final_wide (assuming auto-increment ID is assigned in the order of insertion)
                ids = ids.iloc[::-1].reset_index(drop=True)
                
                # add response_id to wide table
                final_wide["response_id"] = ids["response_id"].values
            else:
                print("survey_response no response_id associated")
                final_wide["response_id"] = None

            #generate Excel file (based on the wide table with response_id) ==========
            #final_wide.to_excel("survey_wide_table.xlsx", index=False)

            excel_path_wide = "survey_wide_table.xlsx"

            if os.path.exists(excel_path_wide):
            # if the doc already exist：add it to the end of Sheet1 
                with pd.ExcelWriter(excel_path_wide, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    start_row = writer.sheets['Sheet1'].max_row
                    final_wide.to_excel(writer, startrow=start_row, index=False, header=False)
            else:
            # if the doc not exist：create a new one
                final_wide.to_excel(excel_path_wide, index=False)
            
            # build long table based on the wide table with response_id (the response_id in the long table will be automatically matched)
            long_df = build_long_table(final_wide)
            #long_df.to_excel("survey_long_table.xlsx", index=False)
            excel_path_long = "survey_long_table.xlsx"
            if os.path.exists(excel_path_long):
                with pd.ExcelWriter(excel_path_long, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    start_row = writer.sheets['Sheet1'].max_row
                    long_df.to_excel(writer, startrow=start_row, index=False, header=False)
            else:
                long_df.to_excel(excel_path_long, index=False)


            print("\n Wide table generated successfully (with ID and filled metadata).")
            print("Long table generated successfully (with ID).")

            # survey_score
            if not long_df.empty and "response_id" in long_df.columns:
                score_df = long_df[~long_df["question_code"].isin(comment_cols)]
                score_df = score_df[["response_id", "question_code", "score"]].dropna(subset=["response_id"])
                
                if not score_df.empty:
                    score_df.to_sql(
                        "survey_score",
                        engine,
                        if_exists="append",
                        index=False
                    )
                    print(f"survey_score finished {len(score_df)} score records")
                else:
                    print("⚠️ survey_score no valid data to write (missing response_id)")
            else:
                print("⚠️ survey_score no data to write")

    except SQLAlchemyError as e:
        print(f"\n❌ failed to write to database: {str(e)}")
    except ValueError as e:
        print(f"\n❌ data length mismatch: {str(e)}")
    except Exception as e:
        print(f"\n❌ program execution failed: {str(e)}")

if __name__ == "__main__":
    main()